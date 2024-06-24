import openpyxl

# 打开工作簿
workbook = openpyxl.load_workbook(r'C:\Users\EDY\Desktop/tops.xlsx')

# 选择Sheet1
sheet = workbook['Sheet1']

# 查找单元格为"top"的所在列
columns_with_top = []
for column in sheet.iter_cols():
    for cell in column:
        if cell.value == 'top':
            columns_with_top.append(cell.column_letter)

# 查找单元格为"top"的所在行
rows_with_top = []
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == 'top':
            rows_with_top.append(cell.row)

print("所在列：", columns_with_top)
print("所在行：", rows_with_top)

import pandas as pd

df = pd.read_excel(r'C:\Users\EDY\Desktop/tops.xlsx', sheet_name='Sheet1', dtype=str)

# 查找单元格为"top"的所在列
columns_with_top = df.columns[['top'].any(axis=1)]

# 查找单元格为"top"的所在行
rows_with_top = df.index[df.isin(['top']).any(axis=1)]

print("所在列：", columns_with_top)
print("所在行：", rows_with_top)